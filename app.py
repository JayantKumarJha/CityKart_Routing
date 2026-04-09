import streamlit as st
import pandas as pd
import numpy as np
import folium
from streamlit_folium import st_folium
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ortools.constraint_solver import routing_enums_pb2
from ortools.constraint_solver import pywrapcp

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CityKart VRP Solver",
    page_icon="🚚",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500&display=swap');

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
h1, h2, h3 { font-family: 'Syne', sans-serif !important; font-weight: 700; }
.stApp { background: #F5F4F0; }

section[data-testid="stSidebar"] {
    background: #0D1B2A !important;
    border-right: 3px solid #FF5733;
}
section[data-testid="stSidebar"] * { color: #E8E4D9 !important; }
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] .stNumberInput label,
section[data-testid="stSidebar"] .stSlider label {
    color: #FF9A7A !important;
    font-family: 'Syne', sans-serif !important;
    font-size: 13px !important;
    font-weight: 600 !important;
    letter-spacing: 0.06em !important;
    text-transform: uppercase !important;
}
section[data-testid="stSidebar"] .stFileUploader label {
    color: #FF9A7A !important;
    font-family: 'Syne', sans-serif !important;
    font-size: 13px !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
}
div[data-testid="metric-container"] {
    background: #FFFFFF;
    border: 1.5px solid #E0DDD5;
    border-left: 4px solid #FF5733;
    border-radius: 8px;
    padding: 12px 16px;
}
div[data-testid="metric-container"] label {
    font-family: 'Syne', sans-serif !important;
    font-size: 11px !important;
    text-transform: uppercase !important;
    letter-spacing: 0.08em !important;
    color: #888 !important;
}
div[data-testid="metric-container"] [data-testid="stMetricValue"] {
    font-family: 'Syne', sans-serif !important;
    font-size: 24px !important;
    font-weight: 800 !important;
    color: #0D1B2A !important;
}
div[data-testid="stButton"] > button {
    background: #FF5733 !important;
    color: white !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 700 !important;
    font-size: 15px !important;
    letter-spacing: 0.05em !important;
    border: none !important;
    border-radius: 6px !important;
    padding: 12px 28px !important;
    width: 100% !important;
    transition: all 0.2s !important;
}
div[data-testid="stButton"] > button:hover { background: #E8421F !important; }
div[data-testid="stDownloadButton"] > button {
    background: #0D1B2A !important;
    color: white !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 700 !important;
    border: none !important;
    border-radius: 6px !important;
    width: 100% !important;
}
.section-header {
    font-family: 'Syne', sans-serif;
    font-size: 13px;
    font-weight: 700;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    color: #FF5733;
    border-bottom: 1px solid #E0DDD5;
    padding-bottom: 6px;
    margin: 20px 0 12px 0;
}
.route-card {
    background: white;
    border-radius: 8px;
    border: 1px solid #E0DDD5;
    border-left: 4px solid #FF5733;
    padding: 14px 18px;
    margin-bottom: 10px;
}
.route-card-title {
    font-family: 'Syne', sans-serif;
    font-size: 15px;
    font-weight: 700;
    color: #0D1B2A;
    margin-bottom: 6px;
}
.route-card-meta { font-size: 13px; color: #555; line-height: 1.8; }
.route-card-seq  { font-size: 12px; color: #888; margin-top: 6px; word-break: break-word; }
.info-box {
    background: #FFF8F6;
    border: 1px solid #FFCFC4;
    border-radius: 8px;
    padding: 14px 18px;
    margin-bottom: 16px;
    font-size: 13px;
    color: #6B2D1F;
}
.page-title {
    font-family: 'Syne', sans-serif;
    font-size: 36px;
    font-weight: 800;
    color: #0D1B2A;
    letter-spacing: -0.02em;
    line-height: 1.1;
}
.page-subtitle { font-size: 15px; color: #888; margin-top: 4px; }
.depot-badge {
    display: inline-block;
    background: #FF5733;
    color: white;
    font-family: 'Syne', sans-serif;
    font-size: 12px;
    font-weight: 700;
    padding: 4px 12px;
    border-radius: 20px;
    margin-top: 8px;
    letter-spacing: 0.05em;
}
.stProgress > div > div { background-color: #FF5733 !important; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
DEPOT_NAME = "CityKart Depot (Gurugram)"
DEPOT_LAT  = 28.43494
DEPOT_LON  = 76.87850

PALETTE = [
    '#E6194B','#3CB44B','#4363D8','#F58231','#911EB4',
    '#42D4F4','#F032E6','#469990','#9A6324','#800000',
    '#808000','#000075','#FF5733','#006FA6','#A30059',
    '#7A4900','#1CE6FF','#B05B3B','#00998F','#5A0007',
]

def get_color(i):
    return PALETTE[i % len(PALETTE)]


# ─────────────────────────────────────────────────────────────────────────────
# NODE
# ─────────────────────────────────────────────────────────────────────────────
class Node:
    def __init__(self, id, x, y, demand_qty=0, demand_val=0):
        self.id         = id
        self.x          = x      # longitude
        self.y          = y      # latitude
        self.demand_qty = demand_qty
        self.demand_val = demand_val


# ─────────────────────────────────────────────────────────────────────────────
# DISTANCE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def haversine(lat1, lon1, lat2, lon2):
    lat1, lon1, lat2, lon2 = map(np.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = np.sin(dlat/2)**2 + np.cos(lat1)*np.cos(lat2)*np.sin(dlon/2)**2
    return 6371 * 2 * np.arcsin(np.sqrt(a))


def route_distance(route, depot):
    if not route:
        return 0.0
    d = haversine(depot.y, depot.x, route[0].y, route[0].x)
    for i in range(len(route) - 1):
        d += haversine(route[i].y, route[i].x, route[i+1].y, route[i+1].x)
    d += haversine(route[-1].y, route[-1].x, depot.y, depot.x)
    return d


def total_distance(solution, depot):
    return sum(route_distance(r, depot) for r in solution)


def route_load(route):
    return sum(n.demand_qty for n in route), sum(n.demand_val for n in route)


# ─────────────────────────────────────────────────────────────────────────────
# OR-TOOLS CVRP SOLVER
# ─────────────────────────────────────────────────────────────────────────────
def solve_vrp_ortools(depot, customers, criteria, cap_qty, cap_val,
                      max_vehicles, time_limit_sec=60):
    """
    Solves CVRP with Google OR-Tools.
    Node 0 = depot, nodes 1..N = customers.

    Capacity dimensions:
      qty   → single dimension on demand_qty
      value → single dimension on demand_val
      both  → two dimensions enforced simultaneously
    """
    all_nodes = [depot] + customers
    n = len(all_nodes)

    # Distance matrix in integer metres (×10 for sub-km precision)
    SCALE = 10
    dist_matrix = [[0] * n for _ in range(n)]
    for i in range(n):
        for j in range(n):
            if i != j:
                km = haversine(all_nodes[i].y, all_nodes[i].x,
                               all_nodes[j].y, all_nodes[j].x)
                dist_matrix[i][j] = int(km * 1000 * SCALE)

    # OR-Tools setup
    manager = pywrapcp.RoutingIndexManager(n, max_vehicles, 0)
    routing = pywrapcp.RoutingModel(manager)

    def distance_callback(from_idx, to_idx):
        return dist_matrix[manager.IndexToNode(from_idx)][manager.IndexToNode(to_idx)]

    transit_cb = routing.RegisterTransitCallback(distance_callback)
    routing.SetArcCostEvaluatorOfAllVehicles(transit_cb)

    # Capacity dimension(s)
    if criteria in ('qty', 'both'):
        def qty_cb(idx):
            return int(all_nodes[manager.IndexToNode(idx)].demand_qty)
        routing.AddDimensionWithVehicleCapacity(
            routing.RegisterUnaryTransitCallback(qty_cb),
            0, [int(cap_qty)] * max_vehicles, True, 'Cap_Qty'
        )

    if criteria in ('value', 'both'):
        def val_cb(idx):
            return int(all_nodes[manager.IndexToNode(idx)].demand_val)
        routing.AddDimensionWithVehicleCapacity(
            routing.RegisterUnaryTransitCallback(val_cb),
            0, [int(cap_val)] * max_vehicles, True, 'Cap_Val'
        )

    # Search parameters
    params = pywrapcp.DefaultRoutingSearchParameters()
    params.first_solution_strategy = (
        routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC
    )
    params.local_search_metaheuristic = (
        routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
    )
    params.time_limit.seconds = time_limit_sec
    params.log_search = False

    solution = routing.SolveWithParameters(params)
    if not solution:
        return None, float('inf')

    # Extract routes (skip depot index 0)
    routes = []
    for v in range(max_vehicles):
        idx   = routing.Start(v)
        route = []
        while not routing.IsEnd(idx):
            node = manager.IndexToNode(idx)
            if node != 0:
                route.append(customers[node - 1])
            idx = solution.Value(routing.NextVar(idx))
        if route:
            routes.append(route)

    return routes, total_distance(routes, depot)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────
def build_excel(best_solution, depot, criteria, cap_qty, cap_val, best_dist):
    wb   = openpyxl.Workbook()
    NAVY = '0D1B2A'
    LITE = 'FFF8F6'
    thin = Side(style='thin', color='DDDDDD')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        c.fill      = PatternFill('solid', start_color=NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = bdr

    def dat(ws, row, col, val, bg='FFFFFF', center=True, wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name='Calibri', size=9)
        c.fill      = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center' if center else 'left',
                                 vertical='center', wrap_text=wrap)
        c.border    = bdr

    # Sheet 1: Route Summary
    ws1 = wb.active
    ws1.title = 'Route Summary'
    ws1.merge_cells('A1:H1')
    c = ws1['A1']
    c.value = (f'CityKart VRP — OR-Tools  |  Criteria: {criteria.upper()}  |  '
               f'Total Distance: {best_dist:,.1f} km  |  Routes: {len(best_solution)}  |  '
               f'Depot: {depot.id}')
    c.font      = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    c.fill      = PatternFill('solid', start_color=NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 28

    for ci, h in enumerate(['Route','Stops','Total Boxes','Total Value (₹)',
                             'Util Qty (%)','Util Value (%)','Distance (km)','Store Sequence'], 1):
        hdr(ws1, 2, ci, h)
    ws1.row_dimensions[2].height = 28
    for ci, w in enumerate([7,7,13,16,13,14,13,90], 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    for ri, route in enumerate(best_solution):
        xrow = ri + 3
        q, v  = route_load(route)
        dist  = route_distance(route, depot)
        uq = round(100*q/cap_qty,1) if criteria in ('qty','both') and cap_qty > 0 else 'N/A'
        uv = round(100*v/cap_val,1) if criteria in ('value','both') and cap_val > 0 else 'N/A'
        seq = depot.id + ' → ' + ' → '.join(n.id for n in route) + ' → ' + depot.id
        bg  = LITE if ri % 2 == 0 else 'FFFFFF'
        for ci, val in enumerate([ri+1,len(route),int(q),round(v,2),uq,uv,round(dist,2),seq],1):
            dat(ws1, xrow, ci, val, bg=bg, center=(ci < 8), wrap=(ci == 8))
        ws1.row_dimensions[xrow].height = 20

    # Sheet 2: Store Details
    ws2 = wb.create_sheet('Store Details')
    ws2.merge_cells('A1:G1')
    c = ws2['A1']
    c.value     = 'Store-Level Route Assignment'
    c.font      = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    c.fill      = PatternFill('solid', start_color=NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 24
    for ci, h in enumerate(['Route','Stop','Store Name','Latitude','Longitude','Boxes','Value (₹)'],1):
        hdr(ws2, 2, ci, h)
    ws2.row_dimensions[2].height = 26
    for ci, w in enumerate([8,7,36,12,12,12,16],1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    xrow = 3
    for ri, route in enumerate(best_solution):
        bg = LITE if ri % 2 == 0 else 'FFFFFF'
        for si, n in enumerate(route):
            for ci, val in enumerate([ri+1,si+1,n.id,round(n.y,6),round(n.x,6),
                                       int(n.demand_qty),round(n.demand_val,2)],1):
                dat(ws2, xrow, ci, val, bg=bg, center=(ci != 3))
            ws2.row_dimensions[xrow].height = 18
            xrow += 1

    # Sheet 3: Leg Details
    ws3 = wb.create_sheet('Leg Details')
    ws3.merge_cells('A1:H1')
    c = ws3['A1']
    c.value     = 'Route Leg-by-Leg Distance Breakdown'
    c.font      = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    c.fill      = PatternFill('solid', start_color=NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 24
    for ci, h in enumerate(['Route','Leg','From','To','Leg Dist (km)',
                             'Cum Dist (km)','Cum Boxes','Cum Value (₹)'],1):
        hdr(ws3, 2, ci, h)
    ws3.row_dimensions[2].height = 26
    for ci, w in enumerate([7,6,30,30,14,14,14,16],1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    xrow = 3
    for ri, route in enumerate(best_solution):
        bg = LITE if ri % 2 == 0 else 'FFFFFF'
        cum_dist, cum_q, cum_v = 0.0, 0, 0.0
        all_nd = [depot] + route + [depot]
        for leg in range(len(all_nd) - 1):
            frm = all_nd[leg]
            to  = all_nd[leg + 1]
            ld  = haversine(frm.y, frm.x, to.y, to.x)
            cum_dist += ld
            if 1 <= leg <= len(route):
                cum_q += route[leg-1].demand_qty
                cum_v += route[leg-1].demand_val
            for ci, val in enumerate([ri+1,leg+1,frm.id,to.id,
                                       round(ld,2),round(cum_dist,2),int(cum_q),round(cum_v,2)],1):
                dat(ws3, xrow, ci, val, bg=bg)
            ws3.row_dimensions[xrow].height = 18
            xrow += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# FOLIUM MAP
# ─────────────────────────────────────────────────────────────────────────────
def build_map(best_solution, depot, criteria, cap_qty, cap_val, best_dist):
    m = folium.Map(location=[DEPOT_LAT, DEPOT_LON], zoom_start=5,
                   tiles='CartoDB positron')

    folium.Marker(
        location=[depot.y, depot.x],
        popup=folium.Popup(
            f'<b>🏭 DEPOT</b><br>{depot.id}<br>'
            f'Lat: {depot.y:.5f}<br>Lon: {depot.x:.5f}', max_width=220),
        icon=folium.Icon(color='red', icon='home', prefix='fa')
    ).add_to(m)

    for i, route in enumerate(best_solution):
        color = get_color(i)
        q, v  = route_load(route)
        dist  = route_distance(route, depot)
        lats  = [n.y for n in route]
        lons  = [n.x for n in route]
        c_lat = np.mean(lats)
        c_lon = np.mean(lons)
        max_r = max(haversine(c_lat, c_lon, n.y, n.x) for n in route) if route else 0

        folium.Circle(
            location=[c_lat, c_lon],
            radius=max(max_r * 1000, 8000),
            color=color, fill=True, fill_opacity=0.07, weight=1.2,
            popup=folium.Popup(
                f'<b>Route {i+1}</b><br>Stops: {len(route)}<br>'
                f'Boxes: {int(q):,}<br>Value: ₹{v:,.0f}<br>Dist: {dist:,.1f} km',
                max_width=200)
        ).add_to(m)

        full_path = ([(depot.y, depot.x)]
                     + [(n.y, n.x) for n in route]
                     + [(depot.y, depot.x)])
        folium.PolyLine(
            full_path, color=color, weight=2.5, opacity=0.9,
            tooltip=f'Route {i+1} — {len(route)} stops, {dist:,.0f} km'
        ).add_to(m)

        for k in range(len(full_path) - 1):
            mid_lat = (full_path[k][0] + full_path[k+1][0]) / 2
            mid_lon = (full_path[k][1] + full_path[k+1][1]) / 2
            folium.CircleMarker(
                location=[mid_lat, mid_lon], radius=2,
                color=color, fill=True, fill_opacity=0.5, weight=0
            ).add_to(m)

        for stop_num, n in enumerate(route, 1):
            cap_info = ''
            if criteria in ('qty',   'both'): cap_info += f'Boxes: {int(n.demand_qty):,}<br>'
            if criteria in ('value', 'both'): cap_info += f'Value: ₹{n.demand_val:,.0f}<br>'
            folium.CircleMarker(
                location=[n.y, n.x], radius=6,
                color='white', weight=1.5, fill=True,
                fill_color=color, fill_opacity=1.0,
                popup=folium.Popup(
                    f'<b>Route {i+1} — Stop {stop_num}</b><br>{n.id}<br>{cap_info}',
                    max_width=220),
                tooltip=n.id
            ).add_to(m)

        folium.Marker(
            location=[c_lat, c_lon],
            icon=folium.DivIcon(
                html=(f'<div style="font-size:11px;font-weight:700;color:white;'
                      f'background:{color};border-radius:4px;padding:2px 5px;'
                      f'white-space:nowrap;box-shadow:0 1px 3px rgba(0,0,0,.4);">'
                      f'R{i+1}</div>'),
                icon_size=(32, 20), icon_anchor=(16, 10)
            )
        ).add_to(m)

    legend = (
        f'<div style="position:fixed;bottom:28px;left:28px;z-index:1000;'
        f'background:white;padding:12px 16px;border-radius:8px;'
        f'border:1px solid #ccc;font-family:sans-serif;font-size:12px;'
        f'box-shadow:0 2px 6px rgba(0,0,0,.2);">'
        f'<b>🚚 OR-Tools VRP</b><br>'
        f'Criteria: <b>{criteria.upper()}</b><br>'
        f'Routes: <b>{len(best_solution)}</b><br>'
        f'Total km: <b>{best_dist:,.0f}</b><br>'
    )
    if criteria in ('qty',   'both'): legend += f'Cap/veh (boxes): <b>{cap_qty:,.0f}</b><br>'
    if criteria in ('value', 'both'): legend += f'Cap/veh (₹): <b>{cap_val:,.0f}</b><br>'
    legend += '</div>'
    m.get_root().html.add_child(folium.Element(legend))
    return m


# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding:16px 0 8px 0;">
        <div style="font-family:'Syne',sans-serif;font-size:22px;font-weight:800;
                    color:#FF5733;letter-spacing:-0.02em;">CITYKART</div>
        <div style="font-family:'DM Sans',sans-serif;font-size:11px;
                    color:#6B8FA8;letter-spacing:0.12em;text-transform:uppercase;">
            VRP Solver — OR-Tools</div>
    </div>
    <hr style="border-color:#1E3A52;margin:8px 0 16px 0;">
    """, unsafe_allow_html=True)

    st.markdown('<p class="section-header" style="color:#FF9A7A">📁 Data Upload</p>',
                unsafe_allow_html=True)
    loc_file = st.file_uploader(
        "Locations File",
        type=['xlsx','xls'],
        help="Columns: Store Name | Longitude | Latitude  (Row 1 = headers)"
    )
    stk_file = st.file_uploader(
        "Stock File",
        type=['xlsx','xls'],
        help="Columns: Store Name | Stock in Value | Stock in Qty  (Row 1 = headers)"
    )

    st.markdown('<p class="section-header" style="color:#FF9A7A;margin-top:20px">⚙️ Configuration</p>',
                unsafe_allow_html=True)

    criteria = st.selectbox(
        "Capacity Criteria",
        options=[
            ("📦  Boxes (Qty)",        "qty"),
            ("💰  Stock Value (₹)",    "value"),
            ("📦 + 💰  Both",         "both"),
        ],
        format_func=lambda x: x[0]
    )
    criteria_val = criteria[1]

    cap_qty = 1.0
    cap_val = 1.0
    if criteria_val in ('qty', 'both'):
        cap_qty = st.number_input(
            "Vehicle Capacity (Boxes)",
            min_value=1, max_value=100000, value=500, step=50
        )
    if criteria_val in ('value', 'both'):
        cap_val = st.number_input(
            "Vehicle Capacity (₹ Value)",
            min_value=1000, max_value=100000000, value=500000, step=10000
        )

    max_vehicles = st.number_input(
        "Max Vehicles / Routes",
        min_value=1, max_value=500, value=15, step=1
    )

    time_limit = st.slider(
        "Solver Time Limit (seconds)",
        min_value=10, max_value=300, value=60, step=10,
        help="OR-Tools searches within this time window. More time = better routes."
    )

    st.markdown('<hr style="border-color:#1E3A52;margin:16px 0;">', unsafe_allow_html=True)
    solve_btn = st.button("🚀  SOLVE VRP")

    st.markdown("""
    <div style="margin-top:24px;font-size:11px;color:#4A6A82;line-height:1.8">
        <b style="color:#FF9A7A">Depot (hardcoded)</b><br>
        Gurugram, Haryana<br>
        Lat: 28.43494  Lon: 76.87850<br><br>
        <b style="color:#FF9A7A">Solver</b><br>
        Google OR-Tools CVRP<br>
        First solution: Path Cheapest Arc<br>
        Metaheuristic: Guided Local Search
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PANEL
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div style="padding:8px 0 20px 0;">
    <div class="page-title">Vehicle Routing Problem</div>
    <div class="page-subtitle">
        Powered by Google OR-Tools · Haversine Distance · Guided Local Search
    </div>
    <div class="depot-badge">🏭 Depot: Gurugram (28.43494, 76.87850)</div>
</div>
""", unsafe_allow_html=True)

if not solve_btn:
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""
        <div class="info-box">
            <b>📋 How to use</b><br>
            1. Upload <b>Locations</b> Excel — Store Name | Longitude | Latitude<br>
            2. Upload <b>Stock</b> Excel — Store Name | Stock in Value | Stock in Qty<br>
            3. Set criteria, capacity, vehicles and time limit in sidebar<br>
            4. Click <b>SOLVE VRP</b>
        </div>""", unsafe_allow_html=True)
    with c2:
        st.markdown("""
        <div class="info-box">
            <b>⚡ About OR-Tools</b><br>
            Builds a haversine distance matrix, then uses Google OR-Tools CVRP solver
            with <b>Guided Local Search</b> to find near-optimal routes within your
            chosen time limit. Longer time = better solution quality.
        </div>""", unsafe_allow_html=True)
    st.info("⬅️  Upload both files and configure parameters, then click **SOLVE VRP**.")
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# LOAD DATA
# ─────────────────────────────────────────────────────────────────────────────
if not loc_file or not stk_file:
    st.error("⚠️  Please upload both Excel files before solving.")
    st.stop()

try:
    df_loc = pd.read_excel(loc_file, header=0)
    df_loc.columns = df_loc.columns.str.strip()
    cm = {}
    for col in df_loc.columns:
        cl = col.lower()
        if 'store' in cl or 'name' in cl: cm['name'] = col
        elif 'lon' in cl: cm['lon'] = col
        elif 'lat' in cl: cm['lat'] = col
    df_loc = df_loc.rename(columns={
        cm.get('name', df_loc.columns[0]): 'store_name',
        cm.get('lon',  df_loc.columns[1]): 'lon',
        cm.get('lat',  df_loc.columns[2]): 'lat'
    })
    df_loc = df_loc[['store_name','lon','lat']].copy()
    df_loc['lon'] = pd.to_numeric(df_loc['lon'], errors='coerce')
    df_loc['lat'] = pd.to_numeric(df_loc['lat'], errors='coerce')
    df_loc = df_loc.dropna(subset=['lon','lat']).reset_index(drop=True)
except Exception as e:
    st.error(f"Error reading locations file: {e}")
    st.stop()

try:
    df_stk = pd.read_excel(stk_file, header=0)
    df_stk.columns = df_stk.columns.str.strip()
    cm2 = {}
    for col in df_stk.columns:
        cl = col.lower()
        if 'store' in cl or 'name' in cl: cm2['name'] = col
        elif 'value' in cl or 'val' in cl: cm2['value'] = col
        elif 'qty' in cl or 'box' in cl or 'quant' in cl: cm2['qty'] = col
    df_stk = df_stk.rename(columns={
        cm2.get('name',  df_stk.columns[0]): 'store_name',
        cm2.get('value', df_stk.columns[1]): 'stock_value',
        cm2.get('qty',   df_stk.columns[2]): 'stock_qty'
    })
    df_stk = df_stk[['store_name','stock_value','stock_qty']].copy()
    df_stk['stock_value'] = pd.to_numeric(df_stk['stock_value'], errors='coerce').fillna(0)
    df_stk['stock_qty']   = pd.to_numeric(df_stk['stock_qty'],   errors='coerce').fillna(0)
except Exception as e:
    st.error(f"Error reading stock file: {e}")
    st.stop()

df = pd.merge(df_loc, df_stk, on='store_name', how='left')
df['stock_value'] = df['stock_value'].fillna(0)
df['stock_qty']   = df['stock_qty'].fillna(0)
df = df.reset_index(drop=True)

depot = Node(id=DEPOT_NAME, x=DEPOT_LON, y=DEPOT_LAT)
customers = [
    Node(id=row['store_name'], x=row['lon'], y=row['lat'],
         demand_qty=row['stock_qty'], demand_val=row['stock_value'])
    for _, row in df.iterrows()
]

# Overload warning
overloaded = []
for n in customers:
    if criteria_val in ('qty', 'both') and n.demand_qty > cap_qty:
        overloaded.append(f"{n.id} (boxes: {int(n.demand_qty):,})")
    elif criteria_val == 'value' and n.demand_val > cap_val:
        overloaded.append(f"{n.id} (value: ₹{n.demand_val:,.0f})")
if overloaded:
    st.warning(f"⚠️ {len(overloaded)} store(s) exceed single-vehicle capacity:\n"
               + "  ".join(overloaded[:8]))

st.info(f"📦 **{len(customers)}** stores loaded. Building {len(customers)+1}×{len(customers)+1} "
        f"distance matrix and solving with OR-Tools…")

# ─────────────────────────────────────────────────────────────────────────────
# SOLVE
# ─────────────────────────────────────────────────────────────────────────────
with st.spinner(f"⚙️  OR-Tools solving — time limit: {time_limit}s"):
    best_solution, best_dist_val = solve_vrp_ortools(
        depot, customers,
        criteria=criteria_val,
        cap_qty=cap_qty,
        cap_val=cap_val,
        max_vehicles=max_vehicles,
        time_limit_sec=time_limit
    )

if best_solution is None:
    st.error(
        "❌  OR-Tools found no feasible solution. "
        "Try: increasing Max Vehicles, relaxing capacity limits, or increasing time limit."
    )
    st.stop()

st.success(f"✅  **{len(best_solution)} routes** found — **{best_dist_val:,.1f} km** total")

# ─────────────────────────────────────────────────────────────────────────────
# RESULTS
# ─────────────────────────────────────────────────────────────────────────────
total_q   = sum(route_load(r)[0] for r in best_solution)
total_v   = sum(route_load(r)[1] for r in best_solution)
avg_stops = np.mean([len(r) for r in best_solution])

m1, m2, m3, m4, m5 = st.columns(5)
m1.metric("Routes",         len(best_solution))
m2.metric("Total Distance", f"{best_dist_val:,.0f} km")
m3.metric("Stores Covered", len(customers))
m4.metric("Total Boxes",    f"{int(total_q):,}")
m5.metric("Total Value",    f"₹{total_v:,.0f}")

st.markdown("---")

tab1, tab2, tab3 = st.tabs(["🗺️  Route Map", "📋  Route Details", "📊  Summary Report"])

with tab1:
    st.markdown('<p class="section-header">Interactive Route Map</p>', unsafe_allow_html=True)
    st.caption("Click any marker or circle for stop details. Route labels shown at cluster centres.")
    map_obj = build_map(best_solution, depot, criteria_val, cap_qty, cap_val, best_dist_val)
    st_folium(map_obj, width=None, height=620, returned_objects=[])

with tab2:
    st.markdown('<p class="section-header">Route-by-Route Breakdown</p>', unsafe_allow_html=True)
    for i, route in enumerate(best_solution):
        q, v  = route_load(route)
        dist  = route_distance(route, depot)
        color = get_color(i)
        uq = f"{100*q/cap_qty:.1f}%" if criteria_val in ('qty','both') and cap_qty > 0 else "N/A"
        uv = f"{100*v/cap_val:.1f}%" if criteria_val in ('value','both') and cap_val > 0 else "N/A"
        seq = " → ".join([depot.id] + [n.id for n in route] + [depot.id])
        util_str = ""
        if criteria_val in ('qty',   'both'): util_str += f"Box Util: <b>{uq}</b> &nbsp;|&nbsp; "
        if criteria_val in ('value', 'both'): util_str += f"Value Util: <b>{uv}</b> &nbsp;|&nbsp; "
        st.markdown(f"""
        <div class="route-card" style="border-left-color:{color}">
            <div class="route-card-title">
                <span style="display:inline-block;width:12px;height:12px;
                       background:{color};border-radius:2px;margin-right:6px;
                       vertical-align:middle;"></span>
                Route {i+1} — {len(route)} stops
            </div>
            <div class="route-card-meta">
                Distance: <b>{dist:,.1f} km</b> &nbsp;|&nbsp;
                Boxes: <b>{int(q):,}</b> &nbsp;|&nbsp;
                Value: <b>₹{v:,.0f}</b> &nbsp;|&nbsp; {util_str}
            </div>
            <div class="route-card-seq">📍 {seq}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown('<p class="section-header" style="margin-top:24px">Stop-Level Detail</p>',
                unsafe_allow_html=True)
    rows = []
    for i, route in enumerate(best_solution):
        cum_q, cum_v, cum_d = 0, 0.0, 0.0
        prev = depot
        for si, n in enumerate(route, 1):
            leg   = haversine(prev.y, prev.x, n.y, n.x)
            cum_d += leg
            cum_q += n.demand_qty
            cum_v += n.demand_val
            rows.append({
                'Route': i+1, 'Stop': si, 'Store': n.id,
                'Lat': round(n.y,5), 'Lon': round(n.x,5),
                'Boxes': int(n.demand_qty), 'Value (₹)': round(n.demand_val,0),
                'Leg (km)': round(leg,1), 'Cum Dist (km)': round(cum_d,1),
                'Cum Boxes': int(cum_q), 'Cum Value (₹)': round(cum_v,0)
            })
            prev = n
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

with tab3:
    st.markdown('<p class="section-header">Summary Report</p>', unsafe_allow_html=True)
    summary_rows = []
    for i, route in enumerate(best_solution):
        q, v = route_load(route)
        dist = route_distance(route, depot)
        uq = round(100*q/cap_qty,1) if criteria_val in ('qty','both') and cap_qty > 0 else None
        uv = round(100*v/cap_val,1) if criteria_val in ('value','both') and cap_val > 0 else None
        summary_rows.append({
            'Route': i+1, 'Stops': len(route),
            'Distance (km)': round(dist,1),
            'Total Boxes': int(q), 'Total Value (₹)': round(v,0),
            'Util Qty (%)': uq, 'Util Value (%)': uv,
            'First Stop': route[0].id if route else '',
            'Last Stop':  route[-1].id if route else '',
        })
    st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

    st.markdown('<p class="section-header" style="margin-top:20px">Grand Totals</p>',
                unsafe_allow_html=True)
    gc1, gc2, gc3 = st.columns(3)
    with gc1:
        st.metric("Total Routes",   len(best_solution))
        st.metric("Total Distance", f"{best_dist_val:,.1f} km")
    with gc2:
        st.metric("Total Stores",    len(customers))
        st.metric("Avg Stops/Route", f"{avg_stops:.1f}")
    with gc3:
        st.metric("Total Boxes",  f"{int(total_q):,}")
        st.metric("Total Value",  f"₹{total_v:,.0f}")
    if criteria_val in ('qty', 'both') and cap_qty > 0:
        st.metric("Avg Box Utilisation",
                  f"{100*total_q/(len(best_solution)*cap_qty):.1f}%")
    if criteria_val in ('value', 'both') and cap_val > 0:
        st.metric("Avg Value Utilisation",
                  f"{100*total_v/(len(best_solution)*cap_val):.1f}%")

st.markdown("---")
st.markdown('<p class="section-header">📥 Download Results</p>', unsafe_allow_html=True)
excel_buf = build_excel(best_solution, depot, criteria_val, cap_qty, cap_val, best_dist_val)
st.download_button(
    label="⬇️  Download Route Report (Excel)",
    data=excel_buf,
    file_name="CityKart_VRP_Routes.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
